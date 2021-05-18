from openpyxl import load_workbook
# wb = load_workbook(filename = 'empty_book.xlsx')
# sheet_ranges = wb['range names']
# print(sheet_ranges['D18'].value)

# wb = load_workbook(filename= "empty_book.xlsx")
# sheet_ranges = wb['range names']
# print(sheet_ranges["D12"].value)


wb = load_workbook(filename = "empty_book.xlsx")
sheet_ranges = wb['Test3']
print(sheet_ranges['F1'].value)
for i in range(1,9):
    sheet_ranges