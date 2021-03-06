from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()

dest_filename = 'empty_book.xlsx'

ws1 = wb.active
ws1.title = "range names"

for row in range(1, 40):
    ws1.append(range(600))

ws2 = wb.create_sheet(title="Pi")

ws2['F5'] = 3.14
ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
print(ws3['AA10'].value)

ws4 = wb.create_sheet(title="test")
for i in range(1,11):
    ws4.cell(column=i,row=1).value="用例编号"


ws5 = wb.c
title1 = ("用例编号","用例模块","用例标题","用例级别","测试环境","测试输入","执行操作","预期结果","验证结果","备注")

wb.save(filename = dest_filename)